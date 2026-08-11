#!/usr/bin/env python3
"""Update price-list-data.js from the public Google Sheets workbook.

Existing products:
  Only stock, price and updateDate are updated for variants where
  type == "BOX" and condition == "✨ Shrink".

Name matching / new products:
  ENboxname is used to canonicalize Japanese/English product names before
  matching. Known English aliases are also normalized. Alias duplicates already
  present in the JS are merged. Only truly new canonical products are added.

Images:
  Existing and newly added products are matched by category + item name against
  data/image-url-map.json. A mapped image URL replaces the current image value.
  If no mapping exists, the current image is preserved; new products remain empty.

Non-target variants and all other fields are preserved.
Before writing, products are stably sorted by the updateDate of the
BOX / ✨ Shrink variant, newest first. Products without a target updateDate
are placed last.
"""
from __future__ import annotations

import argparse
import copy
import csv
import json
import re
import sys
import tempfile
import urllib.request
import unicodedata
from collections import Counter, OrderedDict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SHEET_ID = "1IXVH9SGgtwnFd3ni_Rg-scFaNEk7xQJMnvvj7Mdc4rQ"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
TARGET_TYPE = "BOX"
TARGET_CONDITION = "✨ Shrink"

SECTIONS = {
    "Pokémon": {"item": 2, "type": 4, "condition": 5, "stock": 6, "price": 7, "date": 8},
    "One Piece": {"item": 10, "type": 12, "condition": 13, "stock": 14, "price": 15, "date": 16},
}


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--js", default="en/price-list/price-list-data.js")
    p.add_argument("--xlsx", help="Use a local workbook instead of downloading the public sheet")
    p.add_argument("--report", default="price-list-update-report.csv")
    p.add_argument("--image-map", default="data/image-url-map.json")
    return p.parse_args()


def download_xlsx() -> Path:
    request = urllib.request.Request(
        EXPORT_URL,
        headers={"User-Agent": "Mozilla/5.0 GitHub-Actions-Price-Updater/2.0"},
    )
    try:
        with urllib.request.urlopen(request, timeout=60) as response:
            content_type = response.headers.get("Content-Type", "")
            data = response.read()
    except Exception as exc:
        raise RuntimeError(f"Google Sheet download failed: {exc}") from exc

    if len(data) < 1000 or data[:2] != b"PK":
        raise RuntimeError(
            f"Downloaded content is not an XLSX file: content-type={content_type!r}, bytes={len(data)}"
        )

    tmp = tempfile.NamedTemporaryFile(prefix="price-list-", suffix=".xlsx", delete=False)
    tmp.write(data)
    tmp.close()
    return Path(tmp.name)


def load_js(path: Path) -> list[dict[str, Any]]:
    text = path.read_text(encoding="utf-8")
    match = re.fullmatch(
        r"\s*window\.DARUMA_PRICE_GROUPS\s*=\s*(\[.*\])\s*;\s*",
        text,
        flags=re.S,
    )
    if not match:
        raise RuntimeError(f"Could not parse JavaScript data: {path}")
    return json.loads(match.group(1))


def normalize_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def normalize_date(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    raise RuntimeError(f"Unsupported update date value: {value!r}")


def normalize_name(value: Any) -> str:
    """Normalize names for alias matching without changing display spelling."""
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return "".join(ch for ch in text.casefold() if ch.isalnum())


# Known naming variations seen between PRICE_LIST and the website.
# Keys are raw aliases; values are canonical website names.
MANUAL_ALIASES = {
    "Terastal Fest ex": "Terastal Festival ex",
    "Matchless Fighters": "Peerless Fighters",
    "Nihil Zero": "Munikis Zero",
    "Full Metal Force": "Full Metal Wall",
    "Pokémon Card 151": "Pokemon Card 151",
    "Pokemon Card 151": "Pokemon Card 151",
    "Pokémon GO": "Pokemon GO",
    "Movie Special Pack Great Detective Pikachu": "Movie Special Pack Great Detective Pikachu",
    "ムービースペシャルパック「名探偵ピカチュウ」(SMP2)": "Movie Special Pack Great Detective Pikachu",
    "Ultra Sun": "Ultra Sun",
    "拡張パック「ウルトラサン」(SM5S)": "Ultra Sun",
    "Islands Await You": "Islands Await You",
    "拡張パック「キミを待つ島々」(SM2K)": "Islands Await You",
    "Storm Emeralda": "Storm Emeralda",
    "拡張パック「ストームエメラルダ」": "Storm Emeralda",
}


def read_name_map(wb: Any) -> dict[str, str]:
    """Read Japanese -> English mappings from ENboxname and add known aliases."""
    result: dict[str, str] = {}
    if "ENboxname" in wb.sheetnames:
        ws = wb["ENboxname"]
        for row_no in range(1, ws.max_row + 1):
            jp = ws.cell(row_no, 2).value
            en = ws.cell(row_no, 3).value
            if jp not in (None, "") and en not in (None, ""):
                canonical = str(en).strip()
                result[normalize_name(jp)] = canonical
                result[normalize_name(en)] = canonical
    for alias, canonical in MANUAL_ALIASES.items():
        result[normalize_name(alias)] = canonical
        result[normalize_name(canonical)] = canonical
    return result


def canonical_item_name(item: str, name_map: dict[str, str], existing_name_map: dict[str, str]) -> str:
    key = normalize_name(item)
    if key in name_map:
        mapped = name_map[key]
        # Prefer the exact spelling already used by the site when possible.
        return existing_name_map.get(normalize_name(mapped), mapped)
    if key in existing_name_map:
        return existing_name_map[key]
    return item.strip()


def _variant_rank(variant: dict[str, Any]) -> tuple[str, int]:
    """Newest updateDate wins; same date uses the later worksheet row."""
    return (str(variant.get("updateDate") or ""), int(variant.get("_row") or 0))


def read_sheet_products(
    xlsx_path: Path,
    existing_name_map: dict[str, str],
) -> tuple["OrderedDict[tuple[str, str], dict[str, Any]]", dict[str, str]]:
    """Read products and merge Japanese/English aliases into one canonical product."""
    wb = load_workbook(xlsx_path, data_only=True, read_only=False)
    if "Price-List" not in wb.sheetnames:
        raise RuntimeError("Worksheet 'Price-List' was not found")
    ws = wb["Price-List"]
    name_map = read_name_map(wb)

    # Each canonical product keeps one winning variant per type+condition.
    products: "OrderedDict[tuple[str, str], dict[str, Any]]" = OrderedDict()

    for category, cols in SECTIONS.items():
        current_key: tuple[str, str] | None = None
        for row_no in range(6, ws.max_row + 1):
            item_cell = ws.cell(row_no, cols["item"]).value
            if item_cell not in (None, ""):
                raw_item = str(item_cell).strip()
                item = canonical_item_name(raw_item, name_map, existing_name_map)
                current_key = (category, item)
                products.setdefault(current_key, {"variant_map": OrderedDict(), "aliases": []})
                products[current_key]["aliases"].append(raw_item)

            if current_key is None:
                continue

            row_type = str(ws.cell(row_no, cols["type"]).value or "").strip()
            condition = str(ws.cell(row_no, cols["condition"]).value or "").strip()
            if not row_type or not condition:
                continue

            variant = {
                "type": row_type,
                "condition": condition,
                "stock": normalize_value(ws.cell(row_no, cols["stock"]).value),
                "price": normalize_value(ws.cell(row_no, cols["price"]).value),
                "updateDate": normalize_date(ws.cell(row_no, cols["date"]).value),
                "_row": row_no,
            }
            vkey = (row_type, condition)
            old = products[current_key]["variant_map"].get(vkey)
            if old is None or _variant_rank(variant) >= _variant_rank(old):
                products[current_key]["variant_map"][vkey] = variant

    cleaned: "OrderedDict[tuple[str, str], dict[str, Any]]" = OrderedDict()
    for key, info in products.items():
        variants = []
        target_rows = []
        for variant in info["variant_map"].values():
            row_no = variant.pop("_row")
            variants.append(variant)
            if variant["type"] == TARGET_TYPE and variant["condition"] == TARGET_CONDITION:
                target_rows.append({
                    "stock": variant["stock"],
                    "price": variant["price"],
                    "updateDate": variant["updateDate"],
                    "excelRow": row_no,
                })
        if variants:
            cleaned[key] = {
                "variants": variants,
                "target_rows": target_rows,
                "aliases": list(dict.fromkeys(info["aliases"])),
            }
    return cleaned, name_map


def canonicalize_existing_groups(
    groups: list[dict[str, Any]],
    name_map: dict[str, str],
) -> tuple[list[dict[str, Any]], list[str]]:
    """Remove alias duplicates already created by older workflow versions.

    Prefer the group whose item already equals the canonical name. Otherwise keep
    the first group. Variant values will subsequently be refreshed from the sheet.
    """
    existing_spelling = {
        normalize_name(str(g.get("item", ""))): str(g.get("item", ""))
        for g in groups
        if g.get("category") in ("Pokémon", "One Piece")
    }
    kept: "OrderedDict[tuple[str, str], dict[str, Any]]" = OrderedDict()
    removed: list[str] = []
    passthrough: list[dict[str, Any]] = []

    for group in groups:
        category = str(group.get("category", ""))
        if category not in ("Pokémon", "One Piece"):
            passthrough.append(group)
            continue
        raw_item = str(group.get("item", ""))
        canonical = canonical_item_name(raw_item, name_map, existing_spelling)
        key = (category, canonical)
        candidate = copy.deepcopy(group)
        candidate["item"] = canonical
        if key not in kept:
            kept[key] = candidate
            continue

        current = kept[key]
        current_is_canonical = str(current.get("item")) == canonical
        raw_is_canonical = raw_item == canonical
        if raw_is_canonical and not current_is_canonical:
            removed.append(str(current.get("item")))
            kept[key] = candidate
        else:
            removed.append(raw_item)

    result = list(kept.values())
    # Preserve non-product categories at the end exactly as before.
    result.extend(passthrough)
    return result, removed

def insert_new_group(groups: list[dict[str, Any]], group: dict[str, Any]) -> None:
    """Insert a new product at the end of its category block."""
    category = group["category"]
    last_category_index = -1
    for i, existing in enumerate(groups):
        if existing.get("category") == category:
            last_category_index = i
    if last_category_index >= 0:
        groups.insert(last_category_index + 1, group)
    else:
        groups.append(group)


def sort_groups_by_target_update_date(groups: list[dict[str, Any]]) -> None:
    """Stable-sort products by BOX / ✨ Shrink updateDate, newest first.

    Python's sort is stable, so products with the same updateDate retain
    their existing relative order. Products without a target date go last.
    """
    def target_date(group: dict[str, Any]) -> str:
        for variant in group.get("variants", []):
            if (
                variant.get("type") == TARGET_TYPE
                and variant.get("condition") == TARGET_CONDITION
            ):
                return str(variant.get("updateDate") or "")
        return ""

    groups.sort(key=target_date, reverse=True)




def load_image_map(path: Path) -> dict[tuple[str, str], str]:
    """Load category + item -> image URL mappings. Missing file is non-fatal."""
    if not path.exists():
        print(f"Image map not found; image update skipped: {path}")
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    items = payload.get("items", payload) if isinstance(payload, dict) else {}
    result: dict[tuple[str, str], str] = {}
    for raw_key, raw_url in items.items():
        if not isinstance(raw_key, str) or not isinstance(raw_url, str) or not raw_url.strip():
            continue
        if "\t" not in raw_key:
            continue
        category, item = raw_key.split("\t", 1)
        result[(category.strip(), item.strip())] = raw_url.strip()
    return result


def apply_images(
    groups: list[dict[str, Any]],
    image_map: dict[tuple[str, str], str],
    report_rows: list[dict[str, Any]],
    name_map: dict[str, str],
) -> tuple[int, int]:
    """Apply mapped images using canonical/normalized product names."""
    normalized_images: dict[tuple[str, str], str] = {}
    for (category, item), url in image_map.items():
        canonical = name_map.get(normalize_name(item), item)
        normalized_images[(category, normalize_name(canonical))] = url
        normalized_images[(category, normalize_name(item))] = url

    changed = 0
    mapped = 0
    for group in groups:
        category = str(group.get("category", ""))
        item = str(group.get("item", ""))
        url = normalized_images.get((category, normalize_name(item)))
        if not url:
            continue
        mapped += 1
        old = str(group.get("image", "") or "")
        if old != url:
            group["image"] = url
            changed += 1
            report_rows.append({
                "category": category,
                "item": item,
                "status": "image_updated",
                "details": json.dumps({"old": old, "new": url}, ensure_ascii=False),
            })
    return changed, mapped

def write_js(path: Path, groups: list[dict[str, Any]]) -> None:
    path.write_text(
        "window.DARUMA_PRICE_GROUPS = "
        + json.dumps(groups, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )


def main() -> int:
    args = parse_args()
    js_path = Path(args.js)
    report_path = Path(args.report)
    image_map_path = Path(args.image_map)
    xlsx_path = Path(args.xlsx) if args.xlsx else download_xlsx()

    groups = load_js(js_path)
    before_raw = copy.deepcopy(groups)

    existing_name_map = {
        normalize_name(str(g.get("item", ""))): str(g.get("item", ""))
        for g in groups
        if g.get("category") in ("Pokémon", "One Piece")
    }
    sheet_products, name_map = read_sheet_products(xlsx_path, existing_name_map)

    # Clean alias duplicates left by older workflow runs before calculating updates.
    groups, removed_alias_duplicates = canonicalize_existing_groups(groups, name_map)
    before = copy.deepcopy(groups)

    existing_keys = {
        (str(g.get("category", "")), str(g.get("item", ""))) for g in groups
    }
    new_keys = [key for key in sheet_products if key not in existing_keys]

    # Add newly registered products first. Existing product ordering is preserved;
    # each new product is appended to the end of its category block.
    added_count = 0
    for key in new_keys:
        category, item = key
        info = sheet_products[key]
        new_group = {
            "category": category,
            "item": item,
            "image": "",
            "variants": copy.deepcopy(info["variants"]),
        }
        insert_new_group(groups, new_group)
        added_count += 1

    matched_items: set[tuple[str, str]] = set()
    missing_from_sheet: list[tuple[str, str]] = []
    update_count = 0
    report_rows: list[dict[str, Any]] = []
    for item in removed_alias_duplicates:
        report_rows.append({
            "category": "",
            "item": item,
            "status": "alias_duplicate_removed",
            "details": "Merged into canonical product name",
        })

    # Update existing products only. Newly added products already contain the
    # complete sheet variant data and should not be counted as existing updates.
    original_keys = {
        (str(g.get("category", "")), str(g.get("item", ""))) for g in before
    }

    for group in groups:
        key = (str(group.get("category", "")), str(group.get("item", "")))
        info = sheet_products.get(key)

        if key in new_keys:
            report_rows.append({
                "category": key[0],
                "item": key[1],
                "status": "new_added",
                "details": json.dumps({"variants": len(group.get("variants", []))}, ensure_ascii=False),
            })
            continue

        targets = [
            variant
            for variant in group.get("variants", [])
            if variant.get("type") == TARGET_TYPE
            and variant.get("condition") == TARGET_CONDITION
        ]
        if not targets:
            continue

        source_rows = info["target_rows"] if info else []
        if not source_rows:
            missing_from_sheet.append(key)
            report_rows.append({
                "category": key[0], "item": key[1], "status": "missing_in_sheet", "details": ""
            })
            continue

        matched_items.add(key)
        for index, variant in enumerate(targets):
            src = source_rows[min(index, len(source_rows) - 1)]
            old = {field: variant.get(field) for field in ("stock", "price", "updateDate")}
            for field in ("stock", "price", "updateDate"):
                variant[field] = src[field]
            new = {field: variant.get(field) for field in ("stock", "price", "updateDate")}
            update_count += 1
            report_rows.append({
                "category": key[0],
                "item": key[1],
                "status": "updated" if old != new else "unchanged",
                "details": json.dumps({"old": old, "new": new}, ensure_ascii=False),
            })

    # Apply image URLs after new products and price updates are in place.
    image_map = load_image_map(image_map_path)
    image_changed_count, image_mapped_count = apply_images(groups, image_map, report_rows, name_map)

    # Initial website display follows the JS array order. Sort by the
    # BOX / ✨ Shrink updateDate so recently updated products appear first.
    sort_groups_by_target_update_date(groups)

    # Verify target fields for every product against the workbook.
    mismatches: list[str] = []
    for group in groups:
        key = (str(group.get("category", "")), str(group.get("item", "")))
        info = sheet_products.get(key)
        if not info:
            continue
        source_rows = info["target_rows"]
        targets = [
            variant
            for variant in group.get("variants", [])
            if variant.get("type") == TARGET_TYPE
            and variant.get("condition") == TARGET_CONDITION
        ]
        for index, variant in enumerate(targets):
            if not source_rows:
                continue
            src = source_rows[min(index, len(source_rows) - 1)]
            for field in ("stock", "price", "updateDate"):
                if variant.get(field) != src[field]:
                    mismatches.append(f"{key} variant={index} field={field}")

    # Verify newly added products exactly match all registered sheet variants.
    new_variant_mismatches: list[str] = []
    group_map = {
        (str(g.get("category", "")), str(g.get("item", ""))): g for g in groups
    }
    for key in new_keys:
        group = group_map[key]
        expected = sheet_products[key]["variants"]
        # Image validation is handled by apply_images using normalized aliases.
        if group.get("variants") != expected:
            new_variant_mismatches.append(f"new variants differ from sheet: {key}")

    # Verify existing images, product structure and non-target variants did not change.
    protected_changes: list[str] = []
    current_map = {
        (str(g.get("category", "")), str(g.get("item", ""))): g for g in groups
    }
    for old_group in before:
        key = (str(old_group.get("category", "")), str(old_group.get("item", "")))
        new_group = current_map.get(key)
        if not new_group:
            protected_changes.append(f"existing product removed: {key}")
            continue
        if old_group.get("image") != new_group.get("image"):
            if new_group.get("image") not in set(image_map.values()):
                protected_changes.append(f"unexpected image change: {key}")
        if len(old_group.get("variants", [])) != len(new_group.get("variants", [])):
            protected_changes.append(f"variant count changed: {key}")
            continue
        for old_variant, new_variant in zip(old_group.get("variants", []), new_group.get("variants", [])):
            target = (
                new_variant.get("type") == TARGET_TYPE
                and new_variant.get("condition") == TARGET_CONDITION
            )
            if target:
                old_copy = copy.deepcopy(old_variant)
                for field in ("stock", "price", "updateDate"):
                    old_copy[field] = new_variant.get(field)
                if old_copy != new_variant:
                    protected_changes.append(f"protected target field changed: {key}")
            elif old_variant != new_variant:
                protected_changes.append(f"non-target variant changed: {key}")

    duplicate_items = [
        key
        for key, count in Counter(
            (str(g.get("category", "")), str(g.get("item", ""))) for g in groups
        ).items()
        if count > 1
    ]

    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["category", "item", "status", "details"])
        writer.writeheader()
        writer.writerows(report_rows)

    errors: list[str] = []
    if missing_from_sheet:
        errors.append(f"{len(missing_from_sheet)} existing items are missing from the sheet")
    if mismatches:
        errors.append(f"{len(mismatches)} workbook mismatches")
    if new_variant_mismatches:
        errors.append(f"{len(new_variant_mismatches)} new-product mismatches")
    if protected_changes:
        errors.append(f"{len(protected_changes)} protected-field changes")
    if duplicate_items:
        errors.append(f"{len(duplicate_items)} duplicate items")

    print(f"Workbook: {xlsx_path}")
    print(f"Existing matched products: {len(matched_items)}")
    print(f"Existing target variants checked: {update_count}")
    print(f"New products added: {added_count}")
    print(f"Alias duplicates removed: {len(removed_alias_duplicates)}")
    print(f"Image map matches: {image_mapped_count}")
    print(f"Images updated: {image_changed_count}")
    for key in new_keys:
        print(f"  ADDED: {key[0]} / {key[1]}")
    print(f"Existing items missing from sheet: {len(missing_from_sheet)}")
    print(f"Mismatches: {len(mismatches)}")
    print(f"New-product mismatches: {len(new_variant_mismatches)}")
    print(f"Protected changes: {len(protected_changes)}")
    print(f"Report: {report_path}")

    if errors:
        print("Validation failed:", file=sys.stderr)
        for error in errors:
            print(f"- {error}", file=sys.stderr)
        return 1

    write_js(js_path, groups)
    print(f"Updated: {js_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
